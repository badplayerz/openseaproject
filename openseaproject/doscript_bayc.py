import openpyxl

import subprocess
import math
import threading

PROCESS_NUM = 5 # 开启脚本数量
SCRIPT_NAME = 'sniperOS_bayc'


def getExcelNum():
    excel_file = openpyxl.load_workbook('./resource/otherside_bayc_leak.xlsx')
    sheet = excel_file['Sheet2']
    cell = sheet['A' + str(1)]
    print(cell.value)
    print(sheet.max_row)
    num = sheet.max_row
    # num = 20
    for pageNum in range(1, num + 1):
        cell = sheet['A' + str(pageNum)]
        print('7777777:' + str(cell.value))
    exc_num = math.ceil(num / PROCESS_NUM)  # excel内循环数

    return {'num': num, 'exc_num': exc_num}

def runExcelScript():
    excelDic = getExcelNum()    # excel数据，总行数和需循环行数
    threads = []    # 线程池
    """
    平均给每个线程分配excel循环行数
    """
    for i in range(0, PROCESS_NUM):
        start_num = i * excelDic['exc_num'] + 1  # 循环开始行数（从1开始）
        end_num = (i + 1) * excelDic['exc_num'] + 1  # 循环结束行数
        if end_num > excelDic['num']:
            end_num = excelDic['num'] + 1

        commands = ("cd %s\n"
                    "scrapy crawl sniperOS_bayc -a start_page=%s -a end_page=%s\n"%('/Users/sdbean-zlh/PycharmProjects/openseaproject',start_num,end_num))

        t = runThreadSellCommands(commands)
        threads.append(t)
        t.start()

    for i in threads:
        i.join()

"""
多线程执行
"""
class runThreadSellCommands(threading.Thread):
    def __init__(self,_commands):
        threading.Thread.__init__(self)
        self.commands = _commands
    def run(self):
        shell_result = subprocess.check_output(self.commands, shell=True).decode().strip().split('\n')
        print(self.commands)



if __name__ == "__main__":
    runExcelScript()

